#!/usr/bin/env python3
"""
calculate_metrics.py

Calculates all corpus metrics for both the GBLS corpus (coded article summaries)
and the reference corpus (journal article archive). Writes all output artifacts
and refreshes both metrics explorer JavaScript bundles.

Usage:
    python3 tools/calculate_metrics.py              # both corpora
    python3 tools/calculate_metrics.py --gbls       # GBLS corpus only
    python3 tools/calculate_metrics.py --reference  # reference corpus only

Dependencies: pandas, openpyxl
    pip install pandas openpyxl
"""

import json
import re
import shutil
import sys
from collections import defaultdict
from datetime import datetime, timezone
from itertools import combinations
from pathlib import Path

import pandas as pd

# ── Shared paths ──────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
HUMAN_SOURCES_DIR = PROJECT_ROOT / "0_human_sources"

# ── GBLS corpus paths ─────────────────────────────────────────────────────────
GBLS_DIR = PROJECT_ROOT / "1_coded_gbls_corpus_articles"
GBLS_METRICS_DIR = PROJECT_ROOT / "2_calculated_metrics" / "gbls_corpus_metrics"

GBLS_IDENTITY_COLS = ["article_id", "year", "decade", "citation", "summary_word_count"]

# ── Reference corpus paths ────────────────────────────────────────────────────
ARCHIVE_DIR = PROJECT_ROOT / "1_coded_reference_corpus_articles"
MANIFEST_PATH = ARCHIVE_DIR / "_manifest.jsonl"
REF_METRICS_DIR = PROJECT_ROOT / "2_calculated_metrics" / "reference_corpus_metrics"

# ── Combined JS output (both corpora in one file) ────────────────────────────
# Primary location: new unified GBLS Literature Reviewer app
COMBINED_JS = (
    PROJECT_ROOT / "site" / "public" / "data"
    / "metrics_explorer_data.js"
)
# Legacy locations (for backward compatibility if they exist)
LEGACY_METRICS_EXPLORER = PROJECT_ROOT / "tools" / "metrics_explorer" / "metrics_explorer_data.js"
LEGACY_LIT_CODER = (
    PROJECT_ROOT / "tools" / "web" / "gbls_lit_coder" / "public" / "metrics_explorer"
    / "metrics_explorer_data.js"
)

REF_IDENTITY_COLS = ["article_id", "year", "decade", "journal", "title"]

MANIFEST_FIELD_MAP = {
    "evidence": "Evidence_Type",
    "method": "Primary_Methodology",
    "context": "Library_Context",
    "game": "Game_Format",
    "services": "Service_Area",
    "audience": "Audience",
    "outcomes": "Intended_Outcome",
    "confidence": "Coding_Confidence",
}

FIXED_FIELDS = {
    "Source_Type": "peer_reviewed_journal_article",
    "Peer_Review": "peer_reviewed",
}


# ═════════════════════════════════════════════════════════════════════════════
# SHARED: Schema parsing
# ═════════════════════════════════════════════════════════════════════════════

def find_schema_file(human_sources_dir: Path) -> Path:
    best, best_score = None, 0
    for f in sorted(human_sources_dir.glob("*.md")):
        if f.name.startswith("."):
            continue
        text = f.read_text(encoding="utf-8", errors="replace")
        score = (
            text.count("\n# ")
            + ("lexicon" in f.name.lower()) * 5
            + ("schema" in f.name.lower()) * 5
        )
        if score > best_score:
            best_score, best = score, f
    if best is None:
        raise FileNotFoundError("No schema file found in 0_human_sources/")
    return best


def parse_schema(human_sources_dir: Path) -> tuple[list[str], dict[str, list[str]]]:
    schema_file = find_schema_file(human_sources_dir)
    text = schema_file.read_text(encoding="utf-8", errors="replace")
    field_names: list[str] = []
    field_values: dict[str, list[str]] = {}
    current_field: str | None = None

    for line in text.split("\n"):
        m = re.match(r"^#\s+([A-Za-z_][A-Za-z0-9_]*)\s*:?\s*$", line)
        if m:
            current_field = m.group(1)
            if current_field not in field_names:
                field_names.append(current_field)
                field_values[current_field] = []
            continue
        if current_field is not None:
            m = re.match(r"^-\s+([a-zA-Z][a-zA-Z0-9_]*)", line)
            if m:
                val = m.group(1)
                if val not in field_values[current_field]:
                    field_values[current_field].append(val)

    return field_names, field_values


# ═════════════════════════════════════════════════════════════════════════════
# GBLS CORPUS: source file parsing
# ═════════════════════════════════════════════════════════════════════════════

def _collect_list_items(lines: list[str], start: int) -> tuple[list[str], int]:
    items: list[str] = []
    i = start
    while i < len(lines):
        line = lines[i]
        list_m = re.match(r"^\s*-\s+(.*)", line)
        if list_m:
            item = list_m.group(1).strip()
            if not re.match(r"^[A-Za-z_][A-Za-z0-9_]*\s*:", item):
                items.append(item)
            i += 1
        elif line.strip() == "":
            peek = i + 1
            while peek < len(lines) and lines[peek].strip() == "":
                peek += 1
            if peek < len(lines) and re.match(r"^\s*-\s+", lines[peek]):
                i = peek
            else:
                break
        else:
            break
    return items, i


def _parse_contributions(contrib_text: str) -> list[dict]:
    contributions = []
    raw_blocks = re.split(r"\n(?=\s*-\s+Target_Section\s*:)", contrib_text)
    for block in raw_blocks:
        if not block.strip():
            continue
        block = re.sub(r"^\s*-\s+", "", block, count=1)

        ts_m = re.search(r"^Target_Section\s*:\s*(.+)$", block, re.MULTILINE)
        target_section = ts_m.group(1).strip().strip('"') if ts_m else ""

        ct_m = re.search(
            r"Contribution_Text\s*:\s*>?\s*\n((?:[ \t]+.+\n?)*)",
            block,
        )
        if ct_m:
            contrib_text_val = re.sub(r"\n\s*", " ", ct_m.group(1)).strip()
        else:
            ct_inline = re.search(r"Contribution_Text\s*:\s*(.+)", block)
            contrib_text_val = ct_inline.group(1).strip() if ct_inline else ""

        if contrib_text_val.lower() in (
            "no_direct_addition_recommended",
            "no direct addition recommended",
            "none",
            "",
        ):
            continue

        contributions.append(
            {"target_section": target_section, "contribution_text": contrib_text_val}
        )

    return contributions


def parse_summary_file(path: Path, schema_fields: list[str]) -> dict:
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:
        return {"_error": str(exc), "_path": str(path), "filename": path.name}

    warnings: list[str] = []

    meta_m = re.search(r"^#{1,4}\s+Metadata\s*$", text, re.MULTILINE | re.IGNORECASE)
    summ_m = re.search(r"^#{1,4}\s+Summary\s*$", text, re.MULTILINE | re.IGNORECASE)

    if not meta_m:
        warnings.append("No Metadata heading found")

    if meta_m:
        citation_raw = text[: meta_m.start()].strip()
        meta_end = summ_m.start() if summ_m else len(text)
        meta_section = text[meta_m.end() : meta_end]
        summary_text = text[summ_m.end() :].strip() if summ_m else ""
    else:
        citation_raw = ""
        contrib_m = re.search(r"^Contributions\s*:", text, re.MULTILINE | re.IGNORECASE)
        meta_section = text[: contrib_m.start()] if contrib_m else text
        summary_text = ""

    citation = re.sub(r"^#+\s*", "", citation_raw)
    citation = re.sub(r"^\d+\.\s*", "", citation).strip()

    contrib_start_m = re.search(
        r"^Contributions\s*:", meta_section, re.MULTILINE | re.IGNORECASE
    )
    if contrib_start_m:
        raw_meta = meta_section[: contrib_start_m.start()]
        raw_contrib = meta_section[contrib_start_m.end() :]
    else:
        raw_meta = meta_section
        raw_contrib = ""

    raw_fields: dict[str, str | list[str] | None] = {}
    lines = raw_meta.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i]
        kv_m = re.match(r"^([A-Za-z_][A-Za-z0-9_]*)\s*:\s*(.*)", line)
        if kv_m:
            key = kv_m.group(1)
            value = kv_m.group(2).strip()
            if value:
                raw_fields[key] = value
                i += 1
            else:
                items, next_i = _collect_list_items(lines, i + 1)
                raw_fields[key] = items if items else None
                i = next_i
        else:
            i += 1

    contributions = _parse_contributions(raw_contrib) if raw_contrib.strip() else []

    zotero_key = str(raw_fields.get("Zotero_Item_Key") or "").strip()
    filename_stem = path.stem
    if zotero_key:
        article_id = zotero_key
    else:
        key_m = re.search(r"\(([A-Z0-9]{8})\)$", filename_stem)
        article_id = key_m.group(1) if key_m else filename_stem

    year_raw = str(raw_fields.get("Year") or "").strip()
    if year_raw and year_raw.lower() not in ("n.d.", "nd", "none", "null", ""):
        try:
            year = int(year_raw)
            year_status = "known"
            decade = f"{(year // 10) * 10}s"
        except ValueError:
            year = None
            year_status = "not_dated"
            decade = "n.d."
            warnings.append(f"Unparseable year: {year_raw!r}")
    else:
        year = None
        year_status = "not_dated"
        decade = "n.d."

    controlled: dict[str, str | list[str] | None] = {}
    for field in schema_fields:
        val = raw_fields.get(field)
        if isinstance(val, list):
            seen: set[str] = set()
            clean: list[str] = []
            for v in val:
                v = v.strip()
                if v and v not in seen:
                    seen.add(v)
                    clean.append(v)
            controlled[field] = clean if clean else None
        elif val is not None:
            controlled[field] = str(val).strip() or None
        else:
            controlled[field] = None

    summary_body = re.sub(r"^#{1,6}\s+\*\*.*?\*\*\s*\n", "", summary_text, count=1).strip()
    words = summary_body.split() if summary_body else []

    target_sections = list(
        dict.fromkeys(c["target_section"] for c in contributions if c.get("target_section"))
    )

    return {
        "article_id": article_id,
        "filename": path.name,
        "citation": citation,
        "year": year,
        "year_status": year_status,
        "decade": decade,
        "citation_key": str(raw_fields.get("Citation_Key") or ""),
        "zotero_item_key": zotero_key,
        "better_bibtex_citation_key": str(raw_fields.get("Better_BibTeX_Citation_Key") or ""),
        "attachment_key": str(raw_fields.get("Attachment_Key") or ""),
        "controlled": controlled,
        "summary_text": summary_text,
        "summary_word_count": len(words),
        "summary_character_count": len(summary_body),
        "contributions": contributions,
        "contribution_count": len(contributions),
        "target_sections": target_sections,
        "source_path": str(path),
        "_warnings": warnings,
    }


# ═════════════════════════════════════════════════════════════════════════════
# REFERENCE CORPUS: manifest loading
# ═════════════════════════════════════════════════════════════════════════════

def load_ref_records(manifest_path: Path, schema_fields: list[str]) -> list[dict]:
    records = []
    seen_ids: set[str] = set()

    for line in manifest_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        row = json.loads(line)

        filename = str(row.get("filename") or "")
        stem = Path(filename).stem
        key_m = re.search(r"\(([A-F0-9]{8,12})\)$", stem, re.IGNORECASE)
        article_id = key_m.group(1).upper() if key_m else stem

        base_id = article_id
        suffix = 1
        while article_id in seen_ids:
            article_id = f"{base_id}_{suffix}"
            suffix += 1
        seen_ids.add(article_id)

        year_raw = str(row.get("year") or "").strip()
        if year_raw.isdigit():
            year = int(year_raw)
            decade = f"{(year // 10) * 10}s"
        else:
            year = None
            decade = "n.d."

        controlled: dict[str, str | list[str] | None] = {}
        for field, val in FIXED_FIELDS.items():
            if field in schema_fields:
                controlled[field] = val

        for manifest_key, field_name in MANIFEST_FIELD_MAP.items():
            if field_name not in schema_fields:
                continue
            raw = row.get(manifest_key)
            if isinstance(raw, list):
                clean = [str(v).strip() for v in raw if str(v).strip()]
                controlled[field_name] = clean if clean else None
            elif raw is not None:
                v = str(raw).strip()
                controlled[field_name] = v if v else None
            else:
                controlled[field_name] = None

        records.append(
            {
                "article_id": article_id,
                "filename": filename,
                "year": year,
                "decade": decade,
                "journal": str(row.get("journal") or "").strip(),
                "title": str(row.get("title") or "").strip(),
                "doi": str(row.get("doi") or "").strip(),
                "coding_basis": "abstract_only",
                "controlled": controlled,
            }
        )

    return records


# ═════════════════════════════════════════════════════════════════════════════
# SHARED: Field helpers
# ═════════════════════════════════════════════════════════════════════════════

def field_key(field_name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", field_name.lower()).strip("_")


def value_slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def col_name(fkey: str, vslug: str) -> str:
    return f"{fkey}__{vslug}"


def get_values(record: dict, field_name: str) -> list[str]:
    val = record["controlled"].get(field_name)
    if val is None:
        return []
    if isinstance(val, list):
        return [v for v in val if v]
    return [str(val)] if str(val).strip() else []


def detect_multi_label_fields(records: list[dict], schema_fields: list[str]) -> set[str]:
    multi = set()
    for rec in records:
        for field in schema_fields:
            if isinstance(rec["controlled"].get(field), list):
                multi.add(field)
    return multi


# ═════════════════════════════════════════════════════════════════════════════
# SHARED: DataFrame builders
# ═════════════════════════════════════════════════════════════════════════════

def _coerce_year(df: pd.DataFrame) -> pd.DataFrame:
    if "year" in df.columns:
        df["year"] = pd.array(df["year"], dtype=pd.Int64Dtype())
    return df


def build_features_long(records: list[dict], schema_fields: list[str]) -> pd.DataFrame:
    rows = []
    for rec in records:
        for field in schema_fields:
            fkey = field_key(field)
            for v in get_values(rec, field):
                rows.append(
                    {
                        "article_id": rec["article_id"],
                        "year": rec["year"],
                        "feature_group": fkey,
                        "feature_value": v,
                        "feature_column": col_name(fkey, value_slug(v)),
                        "present": 1,
                    }
                )
    return _coerce_year(
        pd.DataFrame(
            rows,
            columns=["article_id", "year", "feature_group", "feature_value", "feature_column", "present"],
        )
    )


def build_feature_counts(records: list[dict], schema_fields: list[str]) -> pd.DataFrame:
    total = len(records)
    counts: dict[tuple[str, str], int] = defaultdict(int)
    for rec in records:
        for field in schema_fields:
            fkey = field_key(field)
            for v in get_values(rec, field):
                counts[(fkey, v)] += 1

    rows = [
        {
            "feature_group": fkey,
            "feature_value": v,
            "article_count": count,
            "total_articles": total,
            "article_pct": round(count / total * 100, 1) if total else 0,
        }
        for (fkey, v), count in counts.items()
    ]
    df = pd.DataFrame(
        rows,
        columns=["feature_group", "feature_value", "article_count", "total_articles", "article_pct"],
    )
    df["rank_within_group"] = (
        df.groupby("feature_group")["article_count"]
        .rank(method="dense", ascending=False)
        .astype(int)
    )
    return df.sort_values(["feature_group", "rank_within_group"]).reset_index(drop=True)


def build_publication_year_counts(records: list[dict]) -> pd.DataFrame:
    total = len(records)
    year_counts: dict[str, int] = defaultdict(int)
    for rec in records:
        label = str(rec["year"]) if rec["year"] is not None else "n.d."
        year_counts[label] += 1

    rows = [
        {
            "year_label": label,
            "article_count": count,
            "article_pct": round(count / total * 100, 1) if total else 0,
        }
        for label, count in sorted(
            year_counts.items(),
            key=lambda x: (x[0] == "n.d.", int(x[0]) if x[0] != "n.d." else 9999),
        )
    ]
    return pd.DataFrame(rows, columns=["year_label", "article_count", "article_pct"])


def build_feature_year_counts(records: list[dict], schema_fields: list[str]) -> pd.DataFrame:
    year_totals: dict[int | None, int] = defaultdict(int)
    for rec in records:
        year_totals[rec["year"]] += 1

    counts: dict[tuple, int] = defaultdict(int)
    for rec in records:
        for field in schema_fields:
            fkey = field_key(field)
            for v in get_values(rec, field):
                counts[(rec["year"], fkey, v)] += 1

    rows = []
    for (year, fkey, v), count in counts.items():
        denom = year_totals[year]
        rows.append(
            {
                "year": year,
                "feature_group": fkey,
                "feature_value": v,
                "article_count": count,
                "total_articles": denom,
                "article_pct_in_year": round(count / denom * 100, 1) if denom else 0,
            }
        )

    return _coerce_year(
        pd.DataFrame(
            rows,
            columns=["year", "feature_group", "feature_value", "article_count", "total_articles", "article_pct_in_year"],
        )
        .sort_values(["year", "feature_group", "article_count"], ascending=[True, True, False])
        .reset_index(drop=True)
    )


def build_feature_cooccurrence(records: list[dict], schema_fields: list[str]) -> pd.DataFrame:
    total = len(records)
    pair_counts: dict[tuple, int] = defaultdict(int)
    for rec in records:
        assignments = []
        for field in schema_fields:
            fkey = field_key(field)
            for v in get_values(rec, field):
                assignments.append((fkey, v))
        for (fa, va), (fb, vb) in combinations(assignments, 2):
            pair = tuple(sorted([(fa, va), (fb, vb)]))
            pair_counts[pair] += 1

    rows = [
        {
            "feature_group_a": fa,
            "feature_value_a": va,
            "feature_group_b": fb,
            "feature_value_b": vb,
            "article_count": count,
            "article_pct": round(count / total * 100, 1) if total else 0,
        }
        for ((fa, va), (fb, vb)), count in pair_counts.items()
        if count >= 2
    ]
    df = pd.DataFrame(
        rows,
        columns=["feature_group_a", "feature_value_a", "feature_group_b", "feature_value_b", "article_count", "article_pct"],
    )
    return df.sort_values("article_count", ascending=False).reset_index(drop=True)


def cleanup_obsolete_files(metrics_dir: Path, current_files: set[str]) -> list[str]:
    removed = []
    stale_patterns = [
        re.compile(r"^[a-z_]+_matrix\.csv$"),
        re.compile(r"^[a-z_]+_counts\.csv$"),
        re.compile(r"^[a-z_]+_x_[a-z_]+_matrix\.csv$"),
    ]
    for f in metrics_dir.glob("*.csv"):
        if f.name in current_files:
            continue
        if any(p.match(f.name) for p in stale_patterns):
            f.unlink()
            removed.append(f.name)
    return removed


# ═════════════════════════════════════════════════════════════════════════════
# GBLS CORPUS: corpus-specific builders
# ═════════════════════════════════════════════════════════════════════════════

def gbls_build_articles_df(
    records: list[dict],
    schema_fields: list[str],
    multi_label_fields: set[str],
    include_summary: bool = True,
) -> pd.DataFrame:
    rows = []
    for rec in records:
        row: dict = {
            "article_id": rec["article_id"],
            "filename": rec["filename"],
            "citation": rec["citation"],
            "year": rec["year"],
            "year_status": rec["year_status"],
            "decade": rec["decade"],
            "summary_word_count": rec["summary_word_count"],
            "summary_character_count": rec["summary_character_count"],
            "contribution_count": rec["contribution_count"],
            "target_sections": "|".join(rec["target_sections"]) if rec["target_sections"] else "none",
        }
        for field in schema_fields:
            vals = get_values(rec, field)
            row[field] = "|".join(vals) if field in multi_label_fields else (vals[0] if vals else "")
        row["source_path"] = rec["source_path"]
        if include_summary:
            row["summary_text"] = rec["summary_text"]
        rows.append(row)

    cols = [
        "article_id", "filename", "citation", "year", "year_status", "decade",
        "summary_word_count", "summary_character_count", "contribution_count",
        "target_sections",
    ] + list(schema_fields) + ["source_path"]
    if include_summary:
        cols.append("summary_text")

    return _coerce_year(pd.DataFrame(rows, columns=cols))


def gbls_build_feature_matrix(
    records: list[dict],
    schema_fields: list[str],
    feature_cols: list[str],
) -> pd.DataFrame:
    assignments: dict[str, set[str]] = defaultdict(set)
    for rec in records:
        for field in schema_fields:
            fkey = field_key(field)
            for v in get_values(rec, field):
                assignments[rec["article_id"]].add(col_name(fkey, value_slug(v)))

    rows = []
    for rec in records:
        row: dict = {
            "article_id": rec["article_id"],
            "year": rec["year"],
            "decade": rec["decade"],
            "citation": rec["citation"],
            "summary_word_count": rec["summary_word_count"],
        }
        for c in feature_cols:
            row[c] = 1 if c in assignments[rec["article_id"]] else 0
        rows.append(row)

    return _coerce_year(pd.DataFrame(rows, columns=GBLS_IDENTITY_COLS + feature_cols))


def gbls_build_field_matrix(
    records: list[dict],
    field: str,
    feature_cols: list[str],
) -> pd.DataFrame:
    fkey = field_key(field)
    assignments: dict[str, set[str]] = defaultdict(set)
    for rec in records:
        for v in get_values(rec, field):
            assignments[rec["article_id"]].add(col_name(fkey, value_slug(v)))

    rows = []
    for rec in records:
        row: dict = {
            "article_id": rec["article_id"],
            "year": rec["year"],
            "decade": rec["decade"],
            "citation": rec["citation"],
            "summary_word_count": rec["summary_word_count"],
        }
        for c in feature_cols:
            row[c] = 1 if c in assignments[rec["article_id"]] else 0
        rows.append(row)

    return _coerce_year(pd.DataFrame(rows, columns=GBLS_IDENTITY_COLS + feature_cols))


def gbls_build_contributions_df(records: list[dict]) -> pd.DataFrame:
    rows = []
    for rec in records:
        for i, contrib in enumerate(rec["contributions"], start=1):
            rows.append(
                {
                    "article_id": rec["article_id"],
                    "contribution_number": i,
                    "target_section": contrib.get("target_section", ""),
                    "contribution_text": contrib.get("contribution_text", ""),
                }
            )
    return pd.DataFrame(
        rows,
        columns=["article_id", "contribution_number", "target_section", "contribution_text"],
    )


def gbls_build_dataset_summary(records: list[dict], features_long: pd.DataFrame) -> dict:
    known_years = [r["year"] for r in records if r["year"] is not None]
    word_counts = [r["summary_word_count"] for r in records]
    return {
        "total_articles": len(records),
        "earliest_year": min(known_years) if known_years else None,
        "latest_year": max(known_years) if known_years else None,
        "unique_feature_labels": int(features_long["feature_value"].nunique()),
        "total_article_feature_assignments": len(features_long),
        "mean_features_per_article": round(len(features_long) / len(records), 3) if records else 0,
        "median_summary_word_count": float(pd.Series(word_counts).median()) if word_counts else 0,
        "articles_with_contributions": sum(1 for r in records if r["contribution_count"] > 0),
        "not_dated_articles": sum(1 for r in records if r["year"] is None),
    }


def gbls_write_excel(
    schema_fields: list[str],
    field_values: dict[str, list[str]],
    articles_core: pd.DataFrame,
    feature_counts: pd.DataFrame,
    pub_year_counts: pd.DataFrame,
    dataset_summary: dict,
    data_dict: pd.DataFrame,
    output_path: Path,
) -> None:
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [skip] openpyxl not installed – skipping Excel workbook")
        return

    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    fill = PatternFill(fill_type="solid", fgColor="D9E1F2")

    def _write_df(ws, df: pd.DataFrame) -> None:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = bold
                    cell.fill = fill
                    cell.alignment = Alignment(wrap_text=True)
        ws.freeze_panes = ws.cell(row=2, column=1)
        ws.auto_filter.ref = ws.dimensions

    def _autofit(ws) -> None:
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0) for cell in col
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.append(["GBLS Corpus Metrics – Dashboard"])
    ws_dash["A1"].font = Font(bold=True, size=14)
    ws_dash.append([])
    ws_dash.append(["Metric", "Value"])
    for k, v in dataset_summary.items():
        ws_dash.append([k.replace("_", " ").title(), v])
    ws_dash.append([])
    ws_dash.append(["Feature Group", "Unique Values", "Total Assignments"])
    for field in schema_fields:
        fkey = field_key(field)
        grp = feature_counts[feature_counts["feature_group"] == fkey]
        ws_dash.append([field, len(grp), int(grp["article_count"].sum())])
    _autofit(ws_dash)

    ws_yr = wb.create_sheet("Year Counts")
    _write_df(ws_yr, pub_year_counts)
    _autofit(ws_yr)

    ws_fc = wb.create_sheet("All Feature Counts")
    _write_df(ws_fc, feature_counts)
    _autofit(ws_fc)

    ws_dd = wb.create_sheet("Data Dictionary")
    _write_df(ws_dd, data_dict)
    _autofit(ws_dd)

    for field in schema_fields:
        ws_f = wb.create_sheet(field[:31])
        fkey = field_key(field)
        grp = feature_counts[feature_counts["feature_group"] == fkey].drop(columns=["feature_group"])
        _write_df(ws_f, grp)
        _autofit(ws_f)

    wb.save(output_path)


def gbls_build_js_payload(
    records: list[dict],
    schema_fields: list[str],
    multi_label_fields: set[str],
    dataset_summary: dict,
    feature_counts: pd.DataFrame,
    pub_year_counts: pd.DataFrame,
    feature_year_counts: pd.DataFrame,
) -> dict:
    def _label(field: str) -> str:
        return field.replace("_", " ").title()

    years_known = sorted({r["year"] for r in records if r["year"] is not None})

    feature_groups = [
        {
            "key": field_key(f),
            "label": _label(f),
            "column": field_key(f),
            "multi_value": f in multi_label_fields,
        }
        for f in schema_fields
    ]

    articles_out = []
    for rec in records:
        art: dict = {
            "article_id": rec["article_id"],
            "year": rec["year"],
            "citation": rec["citation"],
        }
        for field in schema_fields:
            fkey = field_key(field)
            vals = get_values(rec, field)
            art[fkey] = "|".join(vals) if field in multi_label_fields else (vals[0] if vals else "")
        articles_out.append(art)

    return {
        "summary": dataset_summary,
        "years": years_known,
        "featureGroups": feature_groups,
        "articles": articles_out,
        "featureCounts": [
            {
                "feature_group": row["feature_group"],
                "feature_value": row["feature_value"],
                "article_count": row["article_count"],
                "article_pct": row["article_pct"],
            }
            for _, row in feature_counts.iterrows()
        ],
        "publicationYearCounts": [
            {"year_label": row["year_label"], "article_count": row["article_count"]}
            for _, row in pub_year_counts.iterrows()
        ],
        "featureYearCounts": [
            {
                "year": row["year"],
                "feature_group": row["feature_group"],
                "feature_value": row["feature_value"],
                "article_count": row["article_count"],
                "article_pct_in_year": row["article_pct_in_year"],
            }
            for _, row in feature_year_counts.iterrows()
        ],
    }


def gbls_validate(
    source_files: list[Path],
    records: list[dict],
    parse_warnings: list[str],
    articles_df: pd.DataFrame,
    features_long: pd.DataFrame,
    feature_counts: pd.DataFrame,
    schema_fields: list[str],
) -> dict:
    total_src = len(source_files)
    total_parsed = len(records)
    article_ids = [r["article_id"] for r in records]
    unique_ids = len(set(article_ids))
    issues = list(parse_warnings)

    if total_parsed != total_src:
        issues.append(f"Source count ({total_src}) != parsed count ({total_parsed})")

    blank_ids = [i for i in article_ids if not i or not str(i).strip()]
    if blank_ids:
        issues.append(f"{len(blank_ids)} blank article_id values")

    if total_parsed - unique_ids:
        from collections import Counter
        dupes = [k for k, v in Counter(article_ids).items() if v > 1]
        issues.append(f"Duplicate article_ids: {dupes}")

    if features_long["present"].ne(1).any():
        issues.append("features_long.present contains values other than 1")

    if len(articles_df) != total_parsed:
        issues.append(f"Matrix rows ({len(articles_df)}) != article count ({total_parsed})")

    long_sum = int(features_long.groupby("feature_group")["present"].sum().sum())
    counts_sum = int(feature_counts["article_count"].sum())
    if long_sum != counts_sum:
        issues.append(
            f"Sum of feature_counts.article_count ({counts_sum}) != long-table assignments ({long_sum})"
        )

    for f in GBLS_METRICS_DIR.glob("*.csv"):
        try:
            pd.read_csv(f)
        except Exception as exc:
            issues.append(f"{f.name}: {exc}")

    return {
        "source_markdown_count": total_src,
        "parsed_article_count": total_parsed,
        "unique_article_id_count": unique_ids,
        "duplicate_article_id_count": total_parsed - unique_ids,
        "missing_year_count": sum(1 for r in records if r["year"] is None),
        "article_feature_assignment_count": len(features_long),
        "feature_counts_sum": counts_sum,
        "matrix_row_count": len(articles_df),
        "binary_feature_column_count": len(features_long["feature_column"].unique()),
        "parse_warnings": issues,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


# ═════════════════════════════════════════════════════════════════════════════
# REFERENCE CORPUS: corpus-specific builders
# ═════════════════════════════════════════════════════════════════════════════

def ref_build_articles_df(
    records: list[dict],
    schema_fields: list[str],
    multi_label_fields: set[str],
) -> pd.DataFrame:
    rows = []
    for rec in records:
        row: dict = {
            "article_id": rec["article_id"],
            "filename": rec["filename"],
            "year": rec["year"],
            "decade": rec["decade"],
            "journal": rec["journal"],
            "title": rec["title"],
            "doi": rec["doi"],
            "coding_basis": rec["coding_basis"],
        }
        for field in schema_fields:
            vals = get_values(rec, field)
            row[field] = "|".join(vals) if field in multi_label_fields else (vals[0] if vals else "")
        rows.append(row)

    cols = [
        "article_id", "filename", "year", "decade",
        "journal", "title", "doi", "coding_basis",
    ] + list(schema_fields)
    return _coerce_year(pd.DataFrame(rows, columns=cols))


def ref_build_feature_matrix(
    records: list[dict],
    schema_fields: list[str],
    feature_cols: list[str],
) -> pd.DataFrame:
    assignments: dict[str, set[str]] = defaultdict(set)
    for rec in records:
        for field in schema_fields:
            fkey = field_key(field)
            for v in get_values(rec, field):
                assignments[rec["article_id"]].add(col_name(fkey, value_slug(v)))

    rows = []
    for rec in records:
        row: dict = {
            "article_id": rec["article_id"],
            "year": rec["year"],
            "decade": rec["decade"],
            "journal": rec["journal"],
            "title": rec["title"],
        }
        for c in feature_cols:
            row[c] = 1 if c in assignments[rec["article_id"]] else 0
        rows.append(row)

    return _coerce_year(pd.DataFrame(rows, columns=REF_IDENTITY_COLS + feature_cols))


def ref_build_field_matrix(
    records: list[dict],
    field: str,
    feature_cols: list[str],
) -> pd.DataFrame:
    fkey = field_key(field)
    assignments: dict[str, set[str]] = defaultdict(set)
    for rec in records:
        for v in get_values(rec, field):
            assignments[rec["article_id"]].add(col_name(fkey, value_slug(v)))

    rows = []
    for rec in records:
        row: dict = {
            "article_id": rec["article_id"],
            "year": rec["year"],
            "decade": rec["decade"],
            "journal": rec["journal"],
            "title": rec["title"],
        }
        for c in feature_cols:
            row[c] = 1 if c in assignments[rec["article_id"]] else 0
        rows.append(row)

    return _coerce_year(pd.DataFrame(rows, columns=REF_IDENTITY_COLS + feature_cols))


def ref_build_journal_counts(records: list[dict]) -> pd.DataFrame:
    total = len(records)
    counts: dict[str, int] = defaultdict(int)
    for rec in records:
        counts[rec["journal"]] += 1

    rows = [
        {
            "journal": journal,
            "article_count": count,
            "total_articles": total,
            "article_pct": round(count / total * 100, 1) if total else 0,
        }
        for journal, count in counts.items()
    ]
    df = pd.DataFrame(rows, columns=["journal", "article_count", "total_articles", "article_pct"])
    return df.sort_values("article_count", ascending=False).reset_index(drop=True)


def ref_build_dataset_summary(records: list[dict], features_long: pd.DataFrame) -> dict:
    known_years = [r["year"] for r in records if r["year"] is not None]
    journals = {r["journal"] for r in records if r["journal"]}
    return {
        "total_articles": len(records),
        "earliest_year": min(known_years) if known_years else None,
        "latest_year": max(known_years) if known_years else None,
        "unique_journals": len(journals),
        "unique_feature_labels": int(features_long["feature_value"].nunique()),
        "total_article_feature_assignments": len(features_long),
        "mean_features_per_article": round(len(features_long) / len(records), 3) if records else 0,
        "not_dated_articles": sum(1 for r in records if r["year"] is None),
        "coding_basis": "abstract_only",
    }


def ref_write_excel(
    schema_fields: list[str],
    articles_df: pd.DataFrame,
    feature_counts: pd.DataFrame,
    journal_counts: pd.DataFrame,
    pub_year_counts: pd.DataFrame,
    dataset_summary: dict,
    output_path: Path,
) -> None:
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [skip] openpyxl not installed – skipping Excel workbook")
        return

    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    fill = PatternFill(fill_type="solid", fgColor="D9E1F2")

    def _write_df(ws, df: pd.DataFrame) -> None:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = bold
                    cell.fill = fill
                    cell.alignment = Alignment(wrap_text=True)
        ws.freeze_panes = ws.cell(row=2, column=1)
        ws.auto_filter.ref = ws.dimensions

    def _autofit(ws) -> None:
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0) for cell in col
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.append(["GBLS Reference Corpus Metrics – Dashboard"])
    ws_dash["A1"].font = Font(bold=True, size=14)
    ws_dash.append([])
    ws_dash.append(["Metric", "Value"])
    for k, v in dataset_summary.items():
        ws_dash.append([k.replace("_", " ").title(), v])
    ws_dash.append([])
    ws_dash.append(["Feature Group", "Unique Values", "Total Assignments"])
    for field in schema_fields:
        fkey = field_key(field)
        grp = feature_counts[feature_counts["feature_group"] == fkey]
        ws_dash.append([field, len(grp), int(grp["article_count"].sum())])
    _autofit(ws_dash)

    ws_yr = wb.create_sheet("Year Counts")
    _write_df(ws_yr, pub_year_counts)
    _autofit(ws_yr)

    ws_jn = wb.create_sheet("Journal Counts")
    _write_df(ws_jn, journal_counts)
    _autofit(ws_jn)

    ws_fc = wb.create_sheet("All Feature Counts")
    _write_df(ws_fc, feature_counts)
    _autofit(ws_fc)

    for field in schema_fields:
        ws_f = wb.create_sheet(field[:31])
        fkey = field_key(field)
        grp = feature_counts[feature_counts["feature_group"] == fkey].drop(columns=["feature_group"])
        _write_df(ws_f, grp)
        _autofit(ws_f)

    wb.save(output_path)


def ref_build_js_payload(
    records: list[dict],
    schema_fields: list[str],
    multi_label_fields: set[str],
    dataset_summary: dict,
    feature_counts: pd.DataFrame,
    pub_year_counts: pd.DataFrame,
    feature_year_counts: pd.DataFrame,
    journal_counts: pd.DataFrame,
) -> dict:
    def _label(field: str) -> str:
        return field.replace("_", " ").title()

    years_known = sorted({r["year"] for r in records if r["year"] is not None})
    journals = sorted({r["journal"] for r in records if r["journal"]})

    feature_groups = [
        {
            "key": field_key(f),
            "label": _label(f),
            "column": field_key(f),
            "multi_value": f in multi_label_fields,
        }
        for f in schema_fields
    ]

    articles_out = []
    for rec in records:
        art: dict = {
            "article_id": rec["article_id"],
            "year": rec["year"],
            "journal": rec["journal"],
            "title": rec["title"],
            "doi": rec["doi"],
        }
        for field in schema_fields:
            fkey = field_key(field)
            vals = get_values(rec, field)
            art[fkey] = "|".join(vals) if field in multi_label_fields else (vals[0] if vals else "")
        articles_out.append(art)

    return {
        "summary": dataset_summary,
        "years": years_known,
        "journals": journals,
        "featureGroups": feature_groups,
        "articles": articles_out,
        "featureCounts": [
            {
                "feature_group": row["feature_group"],
                "feature_value": row["feature_value"],
                "article_count": row["article_count"],
                "article_pct": row["article_pct"],
            }
            for _, row in feature_counts.iterrows()
        ],
        "publicationYearCounts": [
            {"year_label": row["year_label"], "article_count": row["article_count"]}
            for _, row in pub_year_counts.iterrows()
        ],
        "featureYearCounts": [
            {
                "year": row["year"],
                "feature_group": row["feature_group"],
                "feature_value": row["feature_value"],
                "article_count": row["article_count"],
                "article_pct_in_year": row["article_pct_in_year"],
            }
            for _, row in feature_year_counts.iterrows()
        ],
        "journalCounts": [
            {"journal": row["journal"], "article_count": row["article_count"]}
            for _, row in journal_counts.iterrows()
        ],
    }


def ref_validate(
    manifest_path: Path,
    records: list[dict],
    articles_df: pd.DataFrame,
    features_long: pd.DataFrame,
    feature_counts: pd.DataFrame,
) -> dict:
    manifest_lines = sum(
        1 for line in manifest_path.read_text(encoding="utf-8").splitlines() if line.strip()
    )
    total_parsed = len(records)
    article_ids = [r["article_id"] for r in records]
    unique_ids = len(set(article_ids))
    issues = []

    if total_parsed != manifest_lines:
        issues.append(f"Manifest lines ({manifest_lines}) != parsed records ({total_parsed})")

    blank_ids = [i for i in article_ids if not i]
    if blank_ids:
        issues.append(f"{len(blank_ids)} blank article_id values")

    if features_long["present"].ne(1).any():
        issues.append("features_long.present contains values other than 1")

    if len(articles_df) != total_parsed:
        issues.append(f"Article df rows ({len(articles_df)}) != parsed count ({total_parsed})")

    long_sum = int(features_long["present"].sum())
    counts_sum = int(feature_counts["article_count"].sum())
    if long_sum != counts_sum:
        issues.append(
            f"Sum of feature_counts.article_count ({counts_sum}) != long-table assignments ({long_sum})"
        )

    for f in REF_METRICS_DIR.glob("*.csv"):
        try:
            pd.read_csv(f)
        except Exception as exc:
            issues.append(f"{f.name}: {exc}")

    return {
        "manifest_line_count": manifest_lines,
        "parsed_article_count": total_parsed,
        "unique_article_id_count": unique_ids,
        "duplicate_article_id_count": total_parsed - unique_ids,
        "missing_year_count": sum(1 for r in records if r["year"] is None),
        "article_feature_assignment_count": len(features_long),
        "feature_counts_sum": counts_sum,
        "matrix_row_count": len(articles_df),
        "binary_feature_column_count": int(features_long["feature_column"].nunique()),
        "issues": issues,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


# ═════════════════════════════════════════════════════════════════════════════
# Combined JS writer
# ═════════════════════════════════════════════════════════════════════════════

def _git_commit_hash() -> str:
    import subprocess
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=PROJECT_ROOT, text=True, stderr=subprocess.DEVNULL,
        ).strip()
    except Exception:
        return "unknown"


def write_combined_js(
    gbls_payload: dict | None,
    ref_payload: dict | None,
    output_path: Path,
    meta: dict | None = None,
) -> None:
    parts = []
    if meta is not None:
        parts.append(
            "window.GBLS_META = "
            + json.dumps(meta, ensure_ascii=False)
            + ";"
        )
    if gbls_payload is not None:
        parts.append(
            "window.GBLS_METRICS = "
            + json.dumps(gbls_payload, ensure_ascii=False, default=str)
            + ";"
        )
    if ref_payload is not None:
        parts.append(
            "window.GBLS_REFERENCE_CORPUS = "
            + json.dumps(ref_payload, ensure_ascii=False, separators=(",", ":"), default=str)
            + ";"
        )
    output_path.write_text("\n".join(parts) + "\n", encoding="utf-8")


# ═════════════════════════════════════════════════════════════════════════════
# PIPELINE: GBLS corpus
# ═════════════════════════════════════════════════════════════════════════════

def build_gbls_corpus() -> None:
    print("=" * 60)
    print("GBLS Corpus Metrics Builder")
    print("=" * 60)

    GBLS_METRICS_DIR.mkdir(parents=True, exist_ok=True)

    print("\n[1/9] Parsing schema...")
    schema_fields, field_values = parse_schema(HUMAN_SOURCES_DIR)
    schema_file = find_schema_file(HUMAN_SOURCES_DIR)
    schema_ref = str(schema_file.relative_to(PROJECT_ROOT))
    print(f"  Schema file : {schema_file.name}")
    print(f"  Fields      : {', '.join(schema_fields)}")

    print("\n[2/9] Discovering source files...")
    all_md = sorted(GBLS_DIR.glob("*.md"))
    source_files = [
        f for f in all_md
        if not f.name.startswith(".") and f.name.lower() != "template.md"
    ]
    print(f"  Found {len(source_files)} source files (excluded template.md and hidden files)")

    print("\n[3/9] Parsing source files...")
    records: list[dict] = []
    parse_warnings: list[str] = []
    errors: list[str] = []

    for path in source_files:
        rec = parse_summary_file(path, schema_fields)
        if "_error" in rec:
            errors.append(f"{path.name}: {rec['_error']}")
            print(f"  ERROR: {path.name}: {rec['_error']}", file=sys.stderr)
            continue
        if rec["_warnings"]:
            for w in rec["_warnings"]:
                parse_warnings.append(f"{path.name}: {w}")
        records.append(rec)

    if errors:
        print(f"\n  {len(errors)} files failed to parse – stopping.")
        for e in errors:
            print(f"    {e}", file=sys.stderr)
        sys.exit(1)

    print(f"  Parsed {len(records)} articles ({len(parse_warnings)} warnings)")
    multi_label_fields = detect_multi_label_fields(records, schema_fields)
    print(f"  Multi-label fields: {', '.join(sorted(multi_label_fields)) or '(none)'}")

    print("\n[4/9] Building data structures...")
    features_long = build_features_long(records, schema_fields)
    feature_cols_ordered = sorted(features_long["feature_column"].unique())

    articles_df = gbls_build_articles_df(records, schema_fields, multi_label_fields, include_summary=True)
    articles_core_df = gbls_build_articles_df(records, schema_fields, multi_label_fields, include_summary=False)
    feature_matrix_df = gbls_build_feature_matrix(records, schema_fields, feature_cols_ordered)
    feature_counts_df = build_feature_counts(records, schema_fields)
    pub_year_df = build_publication_year_counts(records)
    feature_year_df = build_feature_year_counts(records, schema_fields)
    cooccurrence_df = build_feature_cooccurrence(records, schema_fields)
    contributions_df = gbls_build_contributions_df(records)
    dataset_summary = gbls_build_dataset_summary(records, features_long)

    print(f"  Articles           : {len(records)}")
    print(f"  Feature assignments: {len(features_long)}")
    print(f"  Feature columns    : {len(feature_cols_ordered)}")

    print("\n[5/9] Writing CSV files...")
    generated_files: list[Path] = []

    def save_csv(df: pd.DataFrame, name: str) -> Path:
        path = GBLS_METRICS_DIR / name
        df.to_csv(path, index=False, encoding="utf-8")
        generated_files.append(path)
        return path

    save_csv(articles_df, "articles.csv")
    save_csv(articles_core_df, "articles_core.csv")
    save_csv(features_long, "article_features_long.csv")
    save_csv(feature_matrix_df, "article_feature_matrix.csv")
    save_csv(contributions_df, "contributions.csv")
    save_csv(feature_counts_df, "feature_counts.csv")
    save_csv(pub_year_df, "publication_year_counts.csv")
    save_csv(feature_year_df, "feature_year_counts.csv")
    save_csv(cooccurrence_df, "feature_cooccurrence.csv")

    for field in schema_fields:
        fkey = field_key(field)
        field_feature_cols = [c for c in feature_cols_ordered if c.startswith(fkey + "__")]
        save_csv(gbls_build_field_matrix(records, field, field_feature_cols), f"{fkey}_matrix.csv")
        grp = feature_counts_df[feature_counts_df["feature_group"] == fkey].copy()
        save_csv(grp, f"{fkey}_counts.csv")

    ds_df = pd.DataFrame([{"metric": k, "value": v} for k, v in dataset_summary.items()])
    save_csv(ds_df, "dataset_summary.csv")

    ds_json_path = GBLS_METRICS_DIR / "dataset_summary.json"
    ds_json_path.write_text(json.dumps(dataset_summary, indent=2, default=str), encoding="utf-8")
    generated_files.append(ds_json_path)

    print(f"  Wrote {len(generated_files)} CSV/JSON files")

    print("\n[6/9] Writing supporting files...")
    dd_rows = [
        ("articles.csv", "Full article table with summary text", "article_id"),
        ("articles_core.csv", "Article table without summary text", "article_id"),
        ("article_features_long.csv", "One row per article-feature assignment (present=1)", "article_id + feature_column"),
        ("article_feature_matrix.csv", "Binary article × feature matrix across all feature groups", "article_id"),
        ("contributions.csv", "One row per extracted proposed contribution", "article_id + contribution_number"),
        ("feature_counts.csv", "Aggregate counts per feature_group + feature_value", "feature_group + feature_value"),
        ("publication_year_counts.csv", "Article counts by publication year", "year_label"),
        ("feature_year_counts.csv", "Feature counts broken down by publication year", "year + feature_group + feature_value"),
        ("feature_cooccurrence.csv", "Unordered pairs of features co-occurring in ≥2 articles", "feature pair"),
        ("dataset_summary.csv", "Headline corpus metrics", "metric"),
        ("dataset_summary.json", "Same as dataset_summary.csv in JSON", "n/a"),
        ("data_dictionary.csv", "This file – documents all generated artifacts", "file"),
        ("feature_datasets_readme.md", "Human-readable guide to the metric files", "n/a"),
        ("validation_report.json", "Parse and consistency validation results", "n/a"),
        ("gbls_feature_overview.xlsx", "Excel workbook with dashboard and per-field sheets", "n/a"),
        ("generated_artifact_manifest.json", "Manifest of all generated files with sizes and timestamp", "n/a"),
    ]
    for field in schema_fields:
        fkey = field_key(field)
        dd_rows.append((f"{fkey}_matrix.csv", f"Binary article × {field} matrix", "article_id"))
        dd_rows.append((f"{fkey}_counts.csv", f"Aggregate counts for {field}", "feature_group + feature_value"))

    data_dict_df = pd.DataFrame(dd_rows, columns=["file", "description", "primary_key"])
    save_csv(data_dict_df, "data_dictionary.csv")

    readme_path = GBLS_METRICS_DIR / "feature_datasets_readme.md"
    readme_path.write_text(
        f"""# GBLS Feature Datasets

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}
Articles: {len(records)} | Feature assignments: {len(features_long)}

## Quick start (Python / Colab)

```python
import pandas as pd

metrics_dir = "2_calculated_metrics/gbls_corpus_metrics"
articles = pd.read_csv(f"{{metrics_dir}}/articles_core.csv")
features = pd.read_csv(f"{{metrics_dir}}/article_features_long.csv")
matrix   = pd.read_csv(f"{{metrics_dir}}/article_feature_matrix.csv")
```

## Key files

| File | Grain | Purpose |
|------|-------|---------|
| `articles_core.csv` | 1 row/article | All metadata fields; pipe-delimited for multi-label |
| `articles.csv` | 1 row/article | Same + full summary text |
| `article_features_long.csv` | 1 row/assignment | Tidy format; filter by `feature_group` |
| `article_feature_matrix.csv` | 1 row/article | Binary columns across all feature groups |
| `feature_counts.csv` | 1 row/value | Article counts and percentages per feature |
| `feature_cooccurrence.csv` | 1 row/pair | Co-occurring feature pairs (≥2 articles) |

## Counting conventions

- Counts represent coded article presence, not prose term frequency.
- Multi-label fields ({', '.join(sorted(multi_label_fields))}) may contribute multiple assignments per article.
- `article_pct` = article_count / total_articles × 100
- `article_pct_in_year` uses same-year article count as denominator.

## Schema fields

{chr(10).join(f'- **{f}**' + (' (multi-label)' if f in multi_label_fields else '') for f in schema_fields)}

## Undated records

{sum(1 for r in records if r['year'] is None)} article(s) have no year and are labelled `n.d.` in year summaries.
""",
        encoding="utf-8",
    )
    generated_files.append(readme_path)

    print("\n[7/9] Validating...")
    report = gbls_validate(
        source_files, records, parse_warnings,
        articles_core_df, features_long, feature_counts_df, schema_fields,
    )
    vr_path = GBLS_METRICS_DIR / "validation_report.json"
    vr_path.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    generated_files.append(vr_path)
    issue_count = len(report["parse_warnings"])
    if issue_count:
        print(f"  {issue_count} warnings/issues:")
        for w in report["parse_warnings"][:10]:
            print(f"    {w}")
        if issue_count > 10:
            print(f"    ... and {issue_count - 10} more (see validation_report.json)")
    else:
        print("  All checks passed.")

    print("\n[8/9] Writing Excel workbook...")
    xlsx_path = GBLS_METRICS_DIR / "gbls_feature_overview.xlsx"
    gbls_write_excel(
        schema_fields, field_values,
        articles_core_df, feature_counts_df, pub_year_df,
        dataset_summary, data_dict_df, xlsx_path,
    )
    if xlsx_path.exists():
        generated_files.append(xlsx_path)
        print(f"  Wrote {xlsx_path.name}")

    print("\n[9/9] Building JS payload...")
    gbls_payload = gbls_build_js_payload(
        records, schema_fields, multi_label_fields,
        dataset_summary, feature_counts_df, pub_year_df, feature_year_df,
    )
    print("  GBLS payload ready.")

    current_names = {f.name for f in generated_files}
    removed = cleanup_obsolete_files(GBLS_METRICS_DIR, current_names)
    if removed:
        print(f"\n  Removed {len(removed)} obsolete files: {', '.join(removed)}")

    manifest = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "schema_reference": schema_ref,
        "total_articles": len(records),
        "files_generated": {
            f.name: f.stat().st_size for f in sorted(generated_files) if f.exists()
        },
    }
    (GBLS_METRICS_DIR / "generated_artifact_manifest.json").write_text(
        json.dumps(manifest, indent=2), encoding="utf-8"
    )

    print("\n" + "=" * 60)
    print("GBLS BUILD COMPLETE")
    print("=" * 60)
    print(f"  Source summaries parsed : {len(records)}")
    print(f"  Unique articles         : {len(set(r['article_id'] for r in records))}")
    print(f"  Feature assignments     : {len(features_long)}")
    print(f"  Unique coded labels     : {features_long['feature_value'].nunique()}")
    print(f"  Undated sources         : {sum(1 for r in records if r['year'] is None)}")
    print(f"  Workbook created        : {xlsx_path.exists()}")
    print(f"  Output folder           : {GBLS_METRICS_DIR}")
    print(f"  Files generated         : {len(generated_files)}")
    print()

    return gbls_payload


# ═════════════════════════════════════════════════════════════════════════════
# PIPELINE: Reference corpus
# ═════════════════════════════════════════════════════════════════════════════

def build_reference_corpus() -> None:
    print("=" * 60)
    print("Reference Corpus Metrics Builder")
    print("=" * 60)

    REF_METRICS_DIR.mkdir(parents=True, exist_ok=True)

    print("\n[1/8] Parsing schema...")
    schema_fields, field_values = parse_schema(HUMAN_SOURCES_DIR)
    schema_file = find_schema_file(HUMAN_SOURCES_DIR)
    schema_ref = str(schema_file.relative_to(PROJECT_ROOT))
    print(f"  Schema file : {schema_file.name}")
    print(f"  Fields      : {', '.join(schema_fields)}")

    print("\n[2/8] Loading manifest...")
    if not MANIFEST_PATH.exists():
        print(f"  ERROR: manifest not found at {MANIFEST_PATH}", file=sys.stderr)
        sys.exit(1)
    records = load_ref_records(MANIFEST_PATH, schema_fields)
    print(f"  Loaded {len(records)} records from {MANIFEST_PATH.name}")
    multi_label_fields = detect_multi_label_fields(records, schema_fields)
    print(f"  Multi-label fields: {', '.join(sorted(multi_label_fields)) or '(none)'}")

    print("\n[3/8] Building data structures...")
    features_long = build_features_long(records, schema_fields)
    feature_cols = sorted(features_long["feature_column"].unique())

    articles_df = ref_build_articles_df(records, schema_fields, multi_label_fields)
    feature_matrix_df = ref_build_feature_matrix(records, schema_fields, feature_cols)
    feature_counts_df = build_feature_counts(records, schema_fields)
    journal_counts_df = ref_build_journal_counts(records)
    pub_year_df = build_publication_year_counts(records)
    feature_year_df = build_feature_year_counts(records, schema_fields)
    cooccurrence_df = build_feature_cooccurrence(records, schema_fields)
    dataset_summary = ref_build_dataset_summary(records, features_long)

    print(f"  Articles           : {len(records)}")
    print(f"  Feature assignments: {len(features_long)}")
    print(f"  Feature columns    : {len(feature_cols)}")
    print(f"  Journals           : {dataset_summary['unique_journals']}")

    print("\n[4/8] Writing CSV files...")
    generated_files: list[Path] = []

    def save_csv(df: pd.DataFrame, name: str) -> Path:
        path = REF_METRICS_DIR / name
        df.to_csv(path, index=False, encoding="utf-8")
        generated_files.append(path)
        return path

    save_csv(articles_df, "articles.csv")
    save_csv(features_long, "article_features_long.csv")
    save_csv(feature_matrix_df, "article_feature_matrix.csv")
    save_csv(feature_counts_df, "feature_counts.csv")
    save_csv(journal_counts_df, "journal_counts.csv")
    save_csv(pub_year_df, "publication_year_counts.csv")
    save_csv(feature_year_df, "feature_year_counts.csv")
    save_csv(cooccurrence_df, "feature_cooccurrence.csv")

    for field in schema_fields:
        fkey = field_key(field)
        field_cols = [c for c in feature_cols if c.startswith(fkey + "__")]
        save_csv(ref_build_field_matrix(records, field, field_cols), f"{fkey}_matrix.csv")
        grp = feature_counts_df[feature_counts_df["feature_group"] == fkey].copy()
        save_csv(grp, f"{fkey}_counts.csv")

    ds_df = pd.DataFrame([{"metric": k, "value": v} for k, v in dataset_summary.items()])
    save_csv(ds_df, "dataset_summary.csv")

    ds_json_path = REF_METRICS_DIR / "dataset_summary.json"
    ds_json_path.write_text(json.dumps(dataset_summary, indent=2, default=str), encoding="utf-8")
    generated_files.append(ds_json_path)

    print(f"  Wrote {len(generated_files)} CSV/JSON files")

    print("\n[5/8] Writing supporting files...")
    dd_rows = [
        ("articles.csv", "One row per reference article with all coded metadata fields", "article_id"),
        ("article_features_long.csv", "One row per article-feature assignment; always present=1", "article_id + feature_column"),
        ("article_feature_matrix.csv", "Binary article × feature matrix across all feature groups", "article_id"),
        ("feature_counts.csv", "Aggregate article counts per feature_group + feature_value", "feature_group + feature_value"),
        ("journal_counts.csv", "Article counts by journal", "journal"),
        ("publication_year_counts.csv", "Article counts by publication year", "year_label"),
        ("feature_year_counts.csv", "Feature counts by publication year", "year + feature_group + feature_value"),
        ("feature_cooccurrence.csv", "Feature pairs co-occurring in ≥2 articles", "feature pair"),
        ("dataset_summary.csv", "Headline corpus metrics", "metric"),
        ("dataset_summary.json", "Same as dataset_summary.csv in JSON", "n/a"),
        ("data_dictionary.csv", "This file", "file"),
        ("validation_report.json", "Parse and consistency check results", "n/a"),
        ("reference_corpus_overview.xlsx", "Excel workbook with dashboard and per-field sheets", "n/a"),
        ("generated_artifact_manifest.json", "All generated files with sizes and timestamp", "n/a"),
    ]
    for field in schema_fields:
        fkey = field_key(field)
        dd_rows.append((f"{fkey}_matrix.csv", f"Binary article × {field} matrix", "article_id"))
        dd_rows.append((f"{fkey}_counts.csv", f"Aggregate counts for {field}", "feature_group + feature_value"))

    dd_df = pd.DataFrame(dd_rows, columns=["file", "description", "primary_key"])
    save_csv(dd_df, "data_dictionary.csv")

    readme_path = REF_METRICS_DIR / "readme.md"
    readme_path.write_text(
        f"""# GBLS Reference Corpus Metrics

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}
Articles: {len(records)} | Journals: {dataset_summary['unique_journals']} | Feature assignments: {len(features_long)}
Coding basis: abstract-only

## Quick start

```python
import pandas as pd
metrics_dir = "2_calculated_metrics/reference_corpus_metrics"
articles = pd.read_csv(f"{{metrics_dir}}/articles.csv")
features = pd.read_csv(f"{{metrics_dir}}/article_features_long.csv")
matrix   = pd.read_csv(f"{{metrics_dir}}/article_feature_matrix.csv")
```

All coding in this corpus is based on article titles and abstracts only.
It is not a substitute for full-text review.
""",
        encoding="utf-8",
    )
    generated_files.append(readme_path)

    print("\n[6/8] Validating...")
    report = ref_validate(MANIFEST_PATH, records, articles_df, features_long, feature_counts_df)
    vr_path = REF_METRICS_DIR / "validation_report.json"
    vr_path.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    generated_files.append(vr_path)
    issue_count = len(report["issues"])
    if issue_count:
        print(f"  {issue_count} issues:")
        for w in report["issues"][:5]:
            print(f"    {w}")
    else:
        print("  All checks passed.")

    print("\n[7/8] Writing Excel workbook...")
    xlsx_path = REF_METRICS_DIR / "reference_corpus_overview.xlsx"
    ref_write_excel(
        schema_fields, articles_df, feature_counts_df,
        journal_counts_df, pub_year_df, dataset_summary, xlsx_path,
    )
    if xlsx_path.exists():
        generated_files.append(xlsx_path)
        print(f"  Wrote {xlsx_path.name}")

    print("\n[8/8] Building JS payload...")
    ref_payload = ref_build_js_payload(
        records, schema_fields, multi_label_fields,
        dataset_summary, feature_counts_df, pub_year_df,
        feature_year_df, journal_counts_df,
    )
    print("  Reference corpus payload ready.")

    current_names = {f.name for f in generated_files}
    removed = cleanup_obsolete_files(REF_METRICS_DIR, current_names)
    if removed:
        print(f"\n  Removed {len(removed)} obsolete files: {', '.join(removed)}")

    manifest = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "schema_reference": schema_ref,
        "manifest_source": str(MANIFEST_PATH.relative_to(PROJECT_ROOT)),
        "total_articles": len(records),
        "files_generated": {
            f.name: f.stat().st_size for f in sorted(generated_files) if f.exists()
        },
    }
    (REF_METRICS_DIR / "generated_artifact_manifest.json").write_text(
        json.dumps(manifest, indent=2), encoding="utf-8"
    )

    print("\n" + "=" * 60)
    print("REFERENCE CORPUS BUILD COMPLETE")
    print("=" * 60)
    print(f"  Articles processed : {len(records)}")
    print(f"  Unique article IDs : {report['unique_article_id_count']}")
    print(f"  Feature assignments: {len(features_long)}")
    print(f"  Unique coded labels: {features_long['feature_value'].nunique()}")
    print(f"  Journals           : {dataset_summary['unique_journals']}")
    print(f"  Undated records    : {report['missing_year_count']}")
    print(f"  Workbook created   : {xlsx_path.exists()}")
    print(f"  Output folder      : {REF_METRICS_DIR}")
    print(f"  Files generated    : {len(generated_files)}")
    print()

    return ref_payload


# ═════════════════════════════════════════════════════════════════════════════
# Entry point
# ═════════════════════════════════════════════════════════════════════════════

def main() -> None:
    run_gbls = "--reference" not in sys.argv
    run_ref = "--gbls" not in sys.argv

    gbls_payload = build_gbls_corpus() if run_gbls else None
    ref_payload = build_reference_corpus() if run_ref else None

    meta = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "git_commit": _git_commit_hash(),
    }

    print("=" * 60)
    print("Writing combined JS bundle")
    print("=" * 60)
    COMBINED_JS.parent.mkdir(parents=True, exist_ok=True)
    write_combined_js(gbls_payload, ref_payload, COMBINED_JS, meta)
    print(f"  Wrote {COMBINED_JS.relative_to(PROJECT_ROOT)}")
    print(f"  Generated at : {meta['generated_at']}")
    print(f"  Git commit   : {meta['git_commit']}")
    
    # Copy to legacy locations for backward compatibility if they exist
    if LEGACY_METRICS_EXPLORER.parent.exists():
        shutil.copy2(COMBINED_JS, LEGACY_METRICS_EXPLORER)
        print(f"  Copied → {LEGACY_METRICS_EXPLORER.relative_to(PROJECT_ROOT)}")
    if LEGACY_LIT_CODER.parent.exists():
        shutil.copy2(COMBINED_JS, LEGACY_LIT_CODER)
        print(f"  Copied → {LEGACY_LIT_CODER.relative_to(PROJECT_ROOT)}")
    print()


if __name__ == "__main__":
    main()
